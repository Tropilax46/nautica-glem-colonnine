"""Genera/rigenera BOM_e_Budget_NauticaGLEM.xlsx — v2.3.

Aggiunte rispetto a v2.2 (correzioni gap di revisione):
- HLK-PM12 (3W, 250mA) sostituito con HLK-10M12 (10W, 833mA): l'alimentatore
  precedente era sottodimensionato di ~2x rispetto al consumo continuo lato 12V
  (bobine contattore + buck 5V per ESP32+PZEM+LED). HLK-10M12 da' 1,7x di margine.
- Aggiunto kit pressacavi IP68 (PG7-PG16): la cassetta Gewiss GW48007 e' IP66
  ma senza pressacavi sugli ingressi cavi degrada a IP20. Indispensabili per
  ambiente marino salino.
- Aggiunto kit resistori 1/4W: include 4,7 kOhm pull-up obbligatoria per il bus
  OneWire del DS18B20, 220 Ohm gate-MOSFET, scorta per altri usi.

v2.2: sfiato membrana + DS18B20 + cassetta 294x152 (gestione termica).
v2.1: colonna "Link acquisto" cliccabile sui due fogli BOM e Infrastruttura.

Uso:
    cd <cartella progetto>
    pip install openpyxl
    python scripts/build_bom_xlsx.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_primary = os.path.join(_root, "BOM_e_Budget_NauticaGLEM.xlsx")
# Se il file primario e' aperto in Excel (lock), salva su _v23.xlsx
try:
    with open(_primary, "ab"):
        OUT = _primary
except PermissionError:
    OUT = os.path.join(_root, "BOM_e_Budget_NauticaGLEM_v23.xlsx")
    print("WARN: file primario in uso, salvo su", os.path.basename(OUT))

HDR_FILL = PatternFill("solid", start_color="0B4F6C")
HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
TITLE_FONT = Font(bold=True, color="0B4F6C", name="Arial", size=14)
INPUT_FONT = Font(color="0000FF", name="Arial", size=11)
LINK_FONT = Font(color="0563C1", underline="single", name="Arial", size=11)
TOTAL_FILL = PatternFill("solid", start_color="EAF3F7")
TOTAL_FONT = Font(bold=True, name="Arial", size=11)
ASSUMP_FILL = PatternFill("solid", start_color="FFFF00")
BODY_FONT = Font(name="Arial", size=11)
NOTE_FONT = Font(italic=True, color="666666", name="Arial", size=10)

EURO = '"€" #,##0.00;("€" #,##0.00);"-"'
INT = '#,##0;(#,##0);"-"'


def set_link(ws, row, col, url, label="Apri"):
    cell = ws.cell(row=row, column=col, value=label)
    if url:
        cell.hyperlink = url
        cell.font = LINK_FONT
    else:
        cell.font = NOTE_FONT
        cell.value = "—"
    return cell


wb = Workbook()

# ============================================================
# 1) BOM colonnina
# ============================================================
ws = wb.active
ws.title = "BOM colonnina"

ws["A1"] = "Bill of Materials - retrofit smart di UNA colonnina (2 prese elettriche + 2 idriche)"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:G1")

ws["A2"] = "Idraulica gia' funzionante: in questo BOM solo la parte elettrica smart. Link per quantita' pilota (~5 kit)."
ws["A2"].font = NOTE_FONT
ws.merge_cells("A2:G2")

headers = ["#", "Componente", "Modello consigliato", "Q.ta'", "Prezzo unit. (EUR)", "Totale (EUR)", "Link acquisto"]
for i, h in enumerate(headers, 1):
    c = ws.cell(row=4, column=i, value=h)
    c.fill, c.font, c.alignment = HDR_FILL, HDR_FONT, Alignment(horizontal="center")

bom = [
    ("ESP32 WiFi con antenna ESTERNA",
     "ESP32-WROOM-32U DevKit DUBEUYEW (kit 3 pz, antenna + pigtail inclusi)",
     1, 7,
     "https://www.amazon.it/DUBEUYEW-ESP32-DevKitC-Sviluppo-ESP32-WROOM-32U-Arduino/dp/B0F28HM99H",
     "Amazon (kit 3 pz)"),
    ("Antenna WiFi 2.4 GHz IP67 (upgrade outdoor opzionale)",
     "omnidirezionale 7 dBi IP67 SMA, range 700-6000 MHz",
     1, 6,
     "https://www.amazon.it/Omnidirezionale-Amplificatore-Impermeabile-700-6000MHz%EF%BC%8CGSM-Monitoraggio/dp/B0DFTN3VQN",
     "Amazon"),
    ("Misuratore energia + TA",
     "PZEM-004T v3 + TA 100A toroidale split-core",
     2, 10,
     "https://www.amazon.it/multimetro-alimentazione-monitoraggio-misuratore-Trasformatore/dp/B0BD4X2V7S",
     "Amazon"),
    ("Contattore 25A 2NO bobina 12 V DC",
     "Finder 22.32.0.012.4320 (2NO, 12 V AC/DC)",
     2, 20,
     "https://it.rs-online.com/web/p/contattori/7892705",
     "RS Components"),
    ("Driver MOSFET per bobina 12 V (logic-level)",
     "IRLZ44N TO-220AB (kit 5 pz) + diodo 1N4007 flyback (R 220Ohm dal kit resistori)",
     2, 2,
     "https://www.amazon.it/IRLZ44N-MOSFET-55V-220AB-transistor/dp/B0851GKFHN",
     "Amazon (5 pz)"),
    ("Alimentatore AC/DC 12 V (bobine + buck logica)",
     "HiLink HLK-10M12 (12 V 10 W, 833 mA) - kit 2 pz - upgrade da HLK-PM12 sottodim.",
     1, 11,
     "https://www.amazon.it/HI-LINK-HLK-10M12-alimentazione-interruttore-intelligente/dp/B07VM99LYC",
     "Amazon (kit 2 pz)"),
    ("Buck step-down 12 V -> 5 V per ESP32+PZEM",
     "MP1584EN regolabile (kit 6 pz)",
     1, 2,
     "https://www.amazon.it/MP1584EN-regolabile-step-down-alimentazione-converter/dp/B072BMBNG1",
     "Amazon (6 pz)"),
    ("LED stato RGB",
     "Aihasd WS2812B 12-LED anello",
     1, 2,
     "https://www.amazon.it/Aihasd-WS2812-Lampada-integrati-WS2812B/dp/B01NBC5CTH",
     "Amazon"),
    ("Sticker QR custom resistente UV/salino",
     "stampa adesivi vinile o alluminio personalizzati",
     2, 1,
     "https://www.vistaprint.it/etichette-adesivi/adesivi/adesivi-con-codice-qr",
     "VistaPrint"),
    ("Pulsante emergenza fungo NC 22 mm IP66",
     "Push button rosso 1NO+1NC IP66 con cassetta",
     1, 8,
     "https://www.amazon.it/Interruttore-Momentaneo-Resistente-Impermeabile-Interruttori/dp/B082Q2C4RB",
     "Amazon"),
    ("Morsettiera + DIN rail + accessori",
     "HERKINDNESS kit DIN 35 mm con morsetti UK-2.5B",
     1, 8,
     "https://www.amazon.it/Elettrici-Morsettiera-Elettrica-Distribuzione-Universale/dp/B07Y7BCRST",
     "Amazon"),
    ("Cavo 2,5 mm2 + fascette + cassetta interna IP66 (294x152x75 mm, dissipazione passiva)",
     "Cavo H07V-K 100 m + cassetta Gewiss IP66 grande GW48007",
     1, 12,
     "https://www.amazon.it/Gewiss-derivazione-connessione-Dimensioni-294x152x75/dp/B07XLTN8WH",
     "Amazon (cassetta)"),
    ("Sfiato a membrana M12 IP69 (ventilazione passiva contro caldo siciliano)",
     "Vent plug INOX + membrana PES (anti-condensa, -40/+100 C)",
     1, 8,
     "https://it.rs-online.com/web/p/ventilazione-per-contenitori/2884609",
     "RS Components"),
    ("Sensore temperatura DS18B20 waterproof (monitor + auto-protezione firmware)",
     "AZDelivery sonda INOX 1 m (kit 2 pz)",
     1, 3,
     "https://www.amazon.it/AZDelivery-DS18B20-mit-Kabel-Parent/dp/B07Z7XNJ9J",
     "Amazon (kit 2 pz)"),
    ("Kit pressacavi IP68 (per ingressi cavi cassetta IP66)",
     "ARPDJK 32 pz misto M12/M16/M20/M25 - 3-15 mm",
     1, 10,
     "https://www.amazon.it/Pressacavi-Impermeabili-ARPDJK-Regolabile-Pressacavo/dp/B07XF88KP9",
     "Amazon (kit 32 pz)"),
    ("Kit resistori 1/4 W (4,7 kOhm pull-up DS18B20, 220 Ohm gate MOSFET, scorta)",
     "BOJACK assortimento 1000 pz, 25 valori, 5%",
     1, 9,
     "https://www.amazon.it/BOJACK-valori-assortimento-resistori-carbonio/dp/B08FD1XVL6",
     "Amazon (1000 pz)"),
]

row = 5
for descr, model, qta, prezzo, url, link_label in bom:
    n = row - 4
    ws.cell(row=row, column=1, value=n).font = BODY_FONT
    ws.cell(row=row, column=2, value=descr).font = BODY_FONT
    ws.cell(row=row, column=3, value=model).font = BODY_FONT
    ws.cell(row=row, column=4, value=qta).font = INPUT_FONT
    pu = ws.cell(row=row, column=5, value=prezzo)
    pu.font = INPUT_FONT
    pu.number_format = EURO
    tot = ws.cell(row=row, column=6, value="=D" + str(row) + "*E" + str(row))
    tot.font = BODY_FONT
    tot.number_format = EURO
    set_link(ws, row, 7, url, link_label)
    row += 1

bom_tot_row = row
ws.cell(row=bom_tot_row, column=2, value="TOTALE materiali per colonnina").font = TOTAL_FONT
ws.cell(row=bom_tot_row, column=2).fill = TOTAL_FILL
ws.cell(row=bom_tot_row, column=6, value="=SUM(F5:F" + str(row - 1) + ")").font = TOTAL_FONT
ws.cell(row=bom_tot_row, column=6).number_format = EURO
ws.cell(row=bom_tot_row, column=6).fill = TOTAL_FILL

ws.cell(row=bom_tot_row + 2, column=1, value="Note:").font = TOTAL_FONT
notes = [
    "Differenziale 30 mA, magnetotermici e cablaggi 230 V sono presenti nelle colonnine esistenti.",
    "Manodopera elettricista non in BOM: gestita separatamente / interna.",
    "L'antenna riga 2 e' un upgrade outdoor IP67. Il kit DUBEUYEW (riga 1) include gia' una antenna 3 dBi + pigtail IPEX-SMA.",
    "ESP32-U + antenna esterna risolvono la schermatura WiFi della cassetta metallica della colonnina.",
    "Contattore con bobina 12 V DC + driver MOSFET = piu' affidabile e con meno componenti del modulo rele' + contattore 230 V AC.",
    "MOSFET IRLZ44N dissipa <1 mW pilotando la bobina (180 mA): NON serve dissipatore.",
    "Sfiato membrana + cassetta 294x152x75 + DS18B20 firmware = gestione termica cassetta sotto sole siciliano.",
    "Alimentatore HLK-10M12 (10W) dimensionato per: 2x bobine contattore (~4,4W) + buck 5V per ESP32+PZEM+LED (~2W) + margine inrush.",
    "Pressacavi obbligatori su tutti gli ingressi cavi della cassetta IP66: senza, l'IP grade reale crolla a IP20.",
    "Resistore 4,7 kOhm tra DATA e VCC del DS18B20: senza pull-up il bus OneWire non funziona.",
    "Per fatturazione kWh certificata MID: sostituire PZEM con Eastron SDM230-MID (+~40 EUR/colonnina).",
    "I link sono ad oggi 14/05/2026: verificare prezzo e disponibilita' prima dell'ordine.",
]
for i, n in enumerate(notes, bom_tot_row + 3):
    ws.cell(row=i, column=2, value="- " + n).font = NOTE_FONT
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=7)

widths = [4, 50, 42, 8, 14, 14, 24]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A5"

# ============================================================
# 2) Infrastruttura molo
# ============================================================
ws2 = wb.create_sheet("Infrastruttura molo")
ws2["A1"] = "Infrastruttura di molo - rete 100% wireless (mesh outdoor)"
ws2["A1"].font = TITLE_FONT
ws2.merge_cells("A1:F1")

ws2["A2"] = "Nessun cablaggio tra access point: copertura mesh WiFi 6, ogni nodo alimentato in 220 V locale."
ws2["A2"].font = NOTE_FONT
ws2.merge_cells("A2:F2")

for i, h in enumerate(["#", "Voce", "Modello", "Q.ta'", "Prezzo (EUR)", "Link acquisto"], 1):
    c = ws2.cell(row=4, column=i, value=h)
    c.fill, c.font = HDR_FILL, HDR_FONT

infra = [
    ("Router LTE industriale (gateway internet)",
     "Teltonika RUT240 EU Standard",
     1, 250,
     "https://www.amazon.it/Teltonika-RUT240-Standard-Package-Ethernet/dp/B07P8S7LJ9",
     "Amazon"),
    ("SIM dati M2M 5 GB/anno",
     "Vodafone Business M2M / 1NCE (contratto annuale)",
     1, 60,
     "https://www.vodafone.it/business/iot/m2m.html",
     "Vodafone Business"),
    ("Mesh AP outdoor WiFi 6 AX3000 IP65",
     "TP-Link Deco X50-Outdoor",
     3, 150,
     "https://www.amazon.it/TP-Link-Deco-X50-Outdoor-AX3000Mbps-Alimentazione/dp/B0BVKZDF98",
     "Amazon"),
    ("Broker MQTT locale + cache offline",
     "Raspberry Pi 4 4 GB Starter Kit (case + alim. + SD 64 GB)",
     1, 90,
     "https://www.amazon.it/Raspberry-Starter-Alimentatore-Alloggiamento-dissipatore/dp/B0D1N3V2FF",
     "Amazon"),
    ("Cassetta tecnica IP66 grande",
     "Gewiss GW48007 294x152x75 mm IP66 con guida DIN",
     1, 60,
     "https://www.amazon.it/Gewiss-derivazione-connessione-Dimensioni-294x152x75/dp/B07XLTN8WH",
     "Amazon"),
    ("UPS 12 V autonomia ~4 h (per router + Pi)",
     "Alphaelettronica KUPSM010 mini-UPS 10000 mAh",
     1, 40,
     "https://www.amazon.it/Alphaelettronica-KUPSM010-continuit%C3%A0-batteria-telecamere/dp/B0C29FYSV5",
     "Amazon"),
    ("Pali / staffe outdoor per Deco (se servono)",
     "supporti palo universali per AP",
     3, 25,
     "https://www.amazon.it/s?k=supporto+palo+access+point+outdoor",
     "Amazon (ricerca)"),
]

r = 5
for descr, modello, qta, prezzo, url, link_label in infra:
    ws2.cell(row=r, column=1, value=r - 4)
    ws2.cell(row=r, column=2, value=descr)
    ws2.cell(row=r, column=3, value=modello)
    q = ws2.cell(row=r, column=4, value=qta)
    q.font = INPUT_FONT
    p = ws2.cell(row=r, column=5, value=prezzo)
    p.font = INPUT_FONT
    p.number_format = EURO
    set_link(ws2, r, 6, url, link_label)
    r += 1

tot_infra_row = r
ws2.cell(row=r, column=2, value="Subtotale infrastruttura").font = TOTAL_FONT
ws2.cell(row=r, column=2).fill = TOTAL_FILL
ws2.cell(row=r, column=5, value="=SUMPRODUCT(D5:D" + str(r - 1) + ",E5:E" + str(r - 1) + ")")
ws2.cell(row=r, column=5).number_format = EURO
ws2.cell(row=r, column=5).font = TOTAL_FONT
ws2.cell(row=r, column=5).fill = TOTAL_FILL

ws2.cell(row=r + 2, column=1, value="Note:").font = TOTAL_FONT
note_infra = [
    "Deco X50-Outdoor: WiFi 6 AX3000, IP65, range outdoor fino a 70 m; un nodo cablato al router, gli altri in mesh.",
    "Coprire 80 colonnine su un molo di ~150 m richiede tipicamente 2-3 nodi a seconda dell'ostruzione.",
    "Il Raspberry Pi sta nella cassetta tecnica con il router e fa da broker MQTT locale + cache offline.",
    "Tutte le colonnine si collegano al WiFi mesh; non serve cablare un cavo dati a ogni colonnina.",
    "Il piano SIM Vodafone Business M2M e' un esempio: confronta con 1NCE, ThingsMobile o altri provider M2M.",
]
for i, n in enumerate(note_infra, r + 3):
    ws2.cell(row=i, column=2, value="- " + n).font = NOTE_FONT
    ws2.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)

for i, w in enumerate([4, 42, 38, 8, 14, 24], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A5"

# ============================================================
# 3) Riepilogo budget
# ============================================================
BOM_TOT_REF = "'BOM colonnina'!F" + str(bom_tot_row)
INFRA_TOT_REF = "'Infrastruttura molo'!E" + str(tot_infra_row)

ws3 = wb.create_sheet("Riepilogo budget")
ws3["A1"] = "Riepilogo budget - Nautica GLEM colonnine smart"
ws3["A1"].font = TITLE_FONT
ws3.merge_cells("A1:D1")

ws3["A3"] = "Parametri (celle blu = modificabili)"
ws3["A3"].font = TOTAL_FONT
ws3.cell(row=4, column=1, value="Numero colonnine")
n_col = ws3.cell(row=4, column=2, value=50)
n_col.font = INPUT_FONT
n_col.fill = ASSUMP_FILL

ws3.cell(row=5, column=1, value="Costo materiali per colonnina (EUR)")
ws3.cell(row=5, column=2, value="=" + BOM_TOT_REF).number_format = EURO

ws3.cell(row=6, column=1, value="Costo infrastruttura totale (EUR)")
ws3.cell(row=6, column=2, value="=" + INFRA_TOT_REF).number_format = EURO

ws3.cell(row=7, column=1, value="Sviluppo software (EUR)")
c_sw = ws3.cell(row=7, column=2, value=21000)
c_sw.font = INPUT_FONT
c_sw.fill = ASSUMP_FILL
c_sw.number_format = EURO

ws3.cell(row=8, column=1, value="Setup VPS + dominio + grafica + collaudi (EUR)")
c_setup = ws3.cell(row=8, column=2, value=500)
c_setup.font = INPUT_FONT
c_setup.fill = ASSUMP_FILL
c_setup.number_format = EURO

ws3["A10"] = "CAPEX - Investimento iniziale"
ws3["A10"].font = TOTAL_FONT
ws3.cell(row=11, column=1, value="Hardware colonnine (n. unita x prezzo)")
ws3.cell(row=11, column=2, value="=B4*B5").number_format = EURO
ws3.cell(row=12, column=1, value="Infrastruttura di molo")
ws3.cell(row=12, column=2, value="=B6").number_format = EURO
ws3.cell(row=13, column=1, value="Sviluppo software")
ws3.cell(row=13, column=2, value="=B7").number_format = EURO
ws3.cell(row=14, column=1, value="Setup vari")
ws3.cell(row=14, column=2, value="=B8").number_format = EURO
ws3.cell(row=15, column=1, value="TOTALE CAPEX").font = TOTAL_FONT
ws3.cell(row=15, column=1).fill = TOTAL_FILL
ws3.cell(row=15, column=2, value="=SUM(B11:B14)").number_format = EURO
ws3.cell(row=15, column=2).font = TOTAL_FONT
ws3.cell(row=15, column=2).fill = TOTAL_FILL

ws3["A17"] = "OPEX mensile"
ws3["A17"].font = TOTAL_FONT
opex = [
    ("VPS Hetzner CX31", 13),
    ("Backup Storage Box", 4),
    ("SIM dati M2M", 5),
    ("Manutenzione hardware (forfait)", 50),
]
r = 18
for descr, prezzo in opex:
    ws3.cell(row=r, column=1, value=descr)
    c = ws3.cell(row=r, column=2, value=prezzo)
    c.font = INPUT_FONT
    c.number_format = EURO
    r += 1
ws3.cell(row=r, column=1, value="TOTALE OPEX/mese").font = TOTAL_FONT
ws3.cell(row=r, column=1).fill = TOTAL_FILL
ws3.cell(row=r, column=2, value="=SUM(B18:B" + str(r - 1) + ")").font = TOTAL_FONT
ws3.cell(row=r, column=2).number_format = EURO
ws3.cell(row=r, column=2).fill = TOTAL_FILL
opex_total_row = r
ws3.cell(row=r + 1, column=1, value="(+) Stripe ~3% sulle ricariche - variabile").font = NOTE_FONT

ws3["A26"] = "Ricavi attesi (modello semplice)"
ws3["A26"].font = TOTAL_FONT
ws3.cell(row=27, column=1, value="kWh medi per presa/mese (mix anno)")
kwh_pp = ws3.cell(row=27, column=2, value=30)
kwh_pp.font = INPUT_FONT
kwh_pp.fill = ASSUMP_FILL
ws3.cell(row=28, column=1, value="Prese per colonnina")
pp = ws3.cell(row=28, column=2, value=2)
pp.font = INPUT_FONT
pp.fill = ASSUMP_FILL
ws3.cell(row=29, column=1, value="Tariffa applicata (EUR/kWh)")
ws3.cell(row=29, column=2, value=0.55).number_format = EURO
ws3.cell(row=29, column=2).font = INPUT_FONT
ws3.cell(row=29, column=2).fill = ASSUMP_FILL
ws3.cell(row=30, column=1, value="Costo energia (EUR/kWh)")
ws3.cell(row=30, column=2, value=0.30).number_format = EURO
ws3.cell(row=30, column=2).font = INPUT_FONT
ws3.cell(row=30, column=2).fill = ASSUMP_FILL

ws3.cell(row=32, column=1, value="kWh erogati/mese")
ws3.cell(row=32, column=2, value="=B4*B28*B27").number_format = INT
ws3.cell(row=33, column=1, value="Ricavi lordi/mese")
ws3.cell(row=33, column=2, value="=B32*B29").number_format = EURO
ws3.cell(row=34, column=1, value="Costo energia/mese")
ws3.cell(row=34, column=2, value="=B32*B30").number_format = EURO
ws3.cell(row=35, column=1, value="Margine energia/mese")
ws3.cell(row=35, column=2, value="=B33-B34").number_format = EURO
ws3.cell(row=36, column=1, value="OPEX/mese (da sopra)")
ws3.cell(row=36, column=2, value="=B" + str(opex_total_row)).number_format = EURO
ws3.cell(row=37, column=1, value="Utile lordo mensile").font = TOTAL_FONT
ws3.cell(row=37, column=1).fill = TOTAL_FILL
ws3.cell(row=37, column=2, value="=B35-B36").number_format = EURO
ws3.cell(row=37, column=2).font = TOTAL_FONT
ws3.cell(row=37, column=2).fill = TOTAL_FILL
ws3.cell(row=38, column=1, value="Utile lordo annuo")
ws3.cell(row=38, column=2, value="=B37*12").number_format = EURO

ws3.cell(row=40, column=1, value="Payback (mesi) - hardware+infra+setup").font = TOTAL_FONT
ws3.cell(row=40, column=2, value='=IFERROR((B11+B12+B14)/B37,"-")').number_format = "0.0"
ws3.cell(row=41, column=1, value="Payback (mesi) - investimento completo").font = TOTAL_FONT
ws3.cell(row=41, column=2, value='=IFERROR(B15/B37,"-")').number_format = "0.0"

for i, w in enumerate([46, 18, 4, 4], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ============================================================
# 4) Sensitivity
# ============================================================
ws4 = wb.create_sheet("Sensitivity")
ws4["A1"] = "Sensitivity sul numero di colonnine"
ws4["A1"].font = TITLE_FONT
ws4.merge_cells("A1:F1")

hdr = ["N. colonnine", "CAPEX hw+infra+setup", "CAPEX totale (incl. SW)",
       "Ricavi mensili lordi", "Utile mensile", "Payback completo (mesi)"]
for i, h in enumerate(hdr, 1):
    c = ws4.cell(row=3, column=i, value=h)
    c.fill, c.font = HDR_FILL, HDR_FONT

scenarios = [20, 30, 50, 80, 100]
r = 4
for n in scenarios:
    ws4.cell(row=r, column=1, value=n).font = INPUT_FONT
    ws4.cell(row=r, column=2,
        value="=A" + str(r) + "*" + BOM_TOT_REF + "+" + INFRA_TOT_REF + "+'Riepilogo budget'!B8"
    ).number_format = EURO
    ws4.cell(row=r, column=3, value="=B" + str(r) + "+'Riepilogo budget'!B7").number_format = EURO
    ws4.cell(row=r, column=4,
        value="=A" + str(r) + "*'Riepilogo budget'!B28*'Riepilogo budget'!B27*'Riepilogo budget'!B29"
    ).number_format = EURO
    ws4.cell(row=r, column=5,
        value="=A" + str(r) + "*'Riepilogo budget'!B28*'Riepilogo budget'!B27*"
              + "('Riepilogo budget'!B29-'Riepilogo budget'!B30)-'Riepilogo budget'!B" + str(opex_total_row)
    ).number_format = EURO
    ws4.cell(row=r, column=6, value='=IFERROR(C' + str(r) + '/E' + str(r) + ',"-")').number_format = "0.0"
    r += 1

for i, w in enumerate([16, 22, 22, 22, 20, 24], 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4.freeze_panes = "A4"

# ============================================================
# 5) Note / changelog
# ============================================================
ws5 = wb.create_sheet("Note")
ws5["A1"] = "Note, fonti e changelog"
ws5["A1"].font = TITLE_FONT
ws5_notes = [
    "v2.3 (14/05/2026, correzioni gap di revisione ingegneria):",
    "  - HLK-PM12 (3W) -> HLK-10M12 (10W): l'alimentatore precedente era sottodimensionato.",
    "  - Aggiunto kit pressacavi IP68: senza, la cassetta IP66 perde la classificazione.",
    "  - Aggiunto kit resistori (4,7 kOhm pull-up DS18B20 obbligatoria + 220 Ohm gate MOSFET).",
    "",
    "v2.2 (14/05/2026, gestione termica):",
    "  - Aggiunto sfiato a membrana M12 IP69 (RS) per ventilazione passiva contro il caldo siciliano.",
    "  - Aggiunto sensore temperatura DS18B20 waterproof - monitor cassetta e auto-protezione firmware.",
    "  - Cassetta interna passata da 196x152 a 294x152 mm per maggior dissipazione passiva.",
    "  - Conferma: il MOSFET IRLZ44N NON necessita di dissipatore (dissipa <1 mW).",
    "",
    "v2.1 (link acquisto):",
    "  - Aggiunta colonna Link acquisto in BOM colonnina e Infrastruttura molo.",
    "  - Mix Amazon.it + RS Components + VistaPrint + Vodafone Business.",
    "  - Quantita' link studiate per pilota ~5 kit.",
    "",
    "v2 (modifiche su feedback Geno):",
    "  - Eliminati dal BOM: magnetotermico, differenziale, manodopera (gia' presenti / gestiti a parte).",
    "  - ESP32-WROOM-32U + antenna esterna IP67 per superare la schermatura della cassetta metallica.",
    "  - Contattore bobina 12 V DC + driver MOSFET (anziche' modulo rele' + contattore 230 V AC).",
    "  - Infrastruttura: rete 100% wireless mesh (TP-Link Deco X50-Outdoor), nessun switch PoE.",
    "  - Aggiunto Raspberry Pi 4 nella cassetta tecnica come broker MQTT locale + cache offline.",
    "",
    "v1:",
    "  - Prezzi indicativi 2026, IVA esclusa.",
    "  - Le colonnine erogano gia' 220V monofase e acqua. Il retrofit smart e' solo elettrico.",
    "  - Per fatturazione kWh certificata MID: sostituire PZEM con Eastron SDM230-MID (+~40 EUR).",
    "  - Verificare bandi PNRR / Marina 4.0 / Transizione 5.0 prima di acquistare.",
    "",
    "Convenzioni colori (standard finanziario):",
    "  - BLU = input modificabile, NERO = formula, sfondo GIALLO = assunzione chiave, link = AZZURRO sottolineato.",
]
for i, n in enumerate(ws5_notes, 3):
    ws5.cell(row=i, column=1, value=n).font = BODY_FONT
ws5.column_dimensions["A"].width = 130

wb.save(OUT)
print("OK", OUT, "bom_tot_row=", bom_tot_row, "infra_tot_row=", tot_infra_row)
